from openpyxl import  load_workbook

# Загружаем файл

book = load_workbook("cars.xlsx")

#Получаем активный лист
page = book.active

# Итерация по строкам

# for row in page.iter_rows(min_row=1,max_col=2,values_only=True):
#     print(f"Автомобиль {row[0]} стоит {row[1]}")

# Итерация по столбцам
# for field in page.iter_cols(min_row=1,max_col=2,values_only=True):
#     # print(f"Автомобиль {field[0]} стоит {field[1]}")
#     print(field)